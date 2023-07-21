import PySimpleGUI as sg
import common
import openpyxl


# read a cell in the spreadsheet and return it's value with some modifiers
def read_cell(raw_value):
    processed_value = raw_value

    # convert zeros and empty strings to None
    if processed_value == '' or processed_value == 0:
        processed_value = None

    # for strings, convert numbers to int, otherwise convert to lowercase
    if isinstance(processed_value, str):
        if processed_value.isnumeric():
            processed_value = int(processed_value)
        else:
            processed_value = processed_value.lower()
    return processed_value


# compare two values, return true if they're different
def compare(value1, value2):
    # convert zeros and empty strings to None
    if value1 == '' or value1 == 0:
        value1 = None
    if value2 == '' or value2 == 0:
        value2 = None

    # for strings, convert numbers to int, otherwise convert to lowercase
    if isinstance(value1, str):
        if value1.isnumeric():
            value1 = int(value1)
        else:
            value1 = value1.lower()
    if isinstance(value2, str):
        if value2.isnumeric():
            value2 = int(value2)
        else:
            value2 = value2.lower()

    # and now that we've sanitised the input, do the comparison
    return value1 != value2


def compare_files(file1, file2):

    logger.clear()
    logger.log(f"Open File One - {file1}")
    wb_file1 = openpyxl.load_workbook(filename=file1, data_only=True)

    logger.log(f"\nOpen File Two - {file2}")
    wb_file2 = openpyxl.load_workbook(filename=file2, data_only=True)

    for sheetName in wb_file1.sheetnames:
        logger.log("\n\nProcessing " + sheetName + ": ")
        difference_found = False
        sheet1 = wb_file1[sheetName]
        if sheetName in wb_file2.sheetnames:
            sheet2 = wb_file2[sheetName]
            max_rows = max([sheet1.max_row, sheet2.max_row])
            max_columns = max([sheet1.max_column, sheet2.max_column])
            for col_num in range(2, max_columns):
                row_num1 = 0
                row_num2 = 0
                while row_num1 < max_rows:
                    row_num1 += 1
                    row_num2 += 1
                    while 17 <= row_num1 <= max_rows and sheet1.cell(row=row_num1, column=col_num).value == '':
                        row_num1 += 1
                    while 17 <= row_num2 <= max_rows and sheet2.cell(row=row_num2, column=col_num).value == '':
                        row_num2 += 1
                    if row_num1 >= max_rows or row_num2 >= max_rows:
                        break
                    value1 = read_cell(sheet1.cell(row=row_num1, column=col_num).value)
                    value2 = read_cell(sheet2.cell(row=row_num2, column=col_num).value)

                    if value1 != value2:
                        print(sheetName + '!' + str('' or value1) + ' - R1: ' + str(
                            row_num1) + ', R2: ' + str(row_num2))
                        logger.log('\n' + str(sheet1.cell(row=1, column=col_num).value or 'MISSING') + ' (' + str(
                            row_num1) + '): ' + str(
                            value1 or '') + ' <> ' + str(
                            value2 or ''))
                        difference_found = True
            if not difference_found:
                logger.log("OK")
        else:
            logger.log(sheetName + " not found in file two")


###################################################################################################################
# MAIN PROGRAM
###################################################################################################################
if __name__ == '__main__':
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text(
        "This program takes two group options files and compares them")
    ],
        [sg.Text("File one: "), sg.Input(), sg.FileBrowse(key="-FILE1-")],
        [sg.Text("File two: "), sg.Input(), sg.FileBrowse(key="-FILE2-")],
        [sg.Button("Submit")],
        [sg.Text("", key='-PROGRESS-')],
        [sg.Multiline(size=(70, 15), key='-TEXT-', autoscroll=True, expand_x=True, expand_y=True)]]

    allowed_extensions = ["xlsx", "xlsm", "xltx", "xltm"]

    # Create the Window
    window = sg.Window('Compare Group Options Files', layout, resizable=True)
    logger = common.Logger(window, "-PROGRESS-", "-TEXT-")
    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Cancel':  # if user closes window or clicks cancel
            break
        elif event == "Submit":
            if values["-FILE1-"] == '' or values["-FILE2-"] == '':
                logger.log('You must select the input files\n')
            elif values["-FILE1-"][-4:] not in allowed_extensions or values["-FILE2-"][-4:] not in allowed_extensions:
                logger.log(
                    'Both files must be excel format - supported formats are: ' + ','.join(allowed_extensions) + '\n')
            else:
                compare_files(values["-FILE1-"], values["-FILE2-"])

    window.close()
